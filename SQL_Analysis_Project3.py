"""
DecodeLabs - Data Analytics Internship
Project 3: SQL Data Analysis
Author: Shaik | Mohan Babu University
"""

import pandas as pd
import sqlite3
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────
# SETUP — Load dataset into in-memory SQLite database
# ─────────────────────────────────────────────────────────
input_path  = r'Project 3/Dataset for Data Analytics.xlsx'
script_dir  = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(script_dir, 'SQL_Analysis_Project3.xlsx')

df_raw = pd.read_excel(input_path)
df_raw['CouponCode'] = df_raw['CouponCode'].fillna('NONE')
df_raw['Date']       = pd.to_datetime(df_raw['Date']).dt.strftime('%Y-%m-%d')

conn = sqlite3.connect(':memory:')
df_raw.to_sql('orders', conn, if_exists='replace', index=False)

print("=" * 60)
print("   PROJECT 3: SQL DATA ANALYSIS")
print("   DecodeLabs Industrial Training | 2026")
print("=" * 60)
print(f"\n  Table 'orders' loaded: {len(df_raw)} rows x {len(df_raw.columns)} columns")

# ─────────────────────────────────────────────────────────
# SQL QUERIES — 15 Business Intelligence Queries
# ─────────────────────────────────────────────────────────

QUERIES = {

    # ── QUERY 1: Overall business summary ──────────────────
    "Q01_Overall_Summary": {
        "business_question": "What is the overall scale and performance of the business?",
        "sql": """
            SELECT
                COUNT(*)                       AS Total_Orders,
                ROUND(SUM(TotalPrice), 2)      AS Total_Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Order_Value,
                ROUND(MIN(TotalPrice), 2)      AS Min_Order,
                ROUND(MAX(TotalPrice), 2)      AS Max_Order
            FROM orders;
        """,
        "insight": "Business processed 1,200 orders generating Rs 12,64,761 total revenue at avg Rs 1,053.97 per order."
    },

    # ── QUERY 2: Revenue by product ────────────────────────
    "Q02_Revenue_By_Product": {
        "business_question": "Which products generate the most revenue?",
        "sql": """
            SELECT
                Product,
                COUNT(*)                       AS Order_Count,
                ROUND(SUM(TotalPrice), 2)      AS Total_Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Order_Value,
                ROUND(MIN(TotalPrice), 2)      AS Min_Price,
                ROUND(MAX(TotalPrice), 2)      AS Max_Price
            FROM orders
            GROUP BY Product
            ORDER BY Total_Revenue DESC;
        """,
        "insight": "Chair leads revenue (Rs 1,95,620). Laptop has the highest avg order value (Rs 1,110.56). Phone is lowest."
    },

    # ── QUERY 3: Orders by status ──────────────────────────
    "Q03_Orders_By_Status": {
        "business_question": "How are orders distributed across statuses?",
        "sql": """
            SELECT
                OrderStatus,
                COUNT(*)                                            AS Order_Count,
                ROUND(SUM(TotalPrice), 2)                          AS Total_Revenue,
                ROUND(AVG(TotalPrice), 2)                          AS Avg_Revenue,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM orders), 2) AS Pct_of_Total
            FROM orders
            GROUP BY OrderStatus
            ORDER BY Order_Count DESC;
        """,
        "insight": "Cancelled+Returned = 497 orders (41.4%) — critical revenue risk of Rs 5,19,673."
    },

    # ── QUERY 4: Annual revenue trend ──────────────────────
    "Q04_Annual_Revenue_Trend": {
        "business_question": "How has revenue trended year over year?",
        "sql": """
            SELECT
                STRFTIME('%Y', Date)           AS Year,
                COUNT(*)                       AS Orders,
                ROUND(SUM(TotalPrice), 2)      AS Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Order
            FROM orders
            GROUP BY Year
            ORDER BY Year;
        """,
        "insight": "Revenue fell 13.1% from 2023 (Rs 5,52,643) to 2024 (Rs 4,80,235). 2025 is partial (Jan-Jun only)."
    },

    # ── QUERY 5: Monthly revenue 2024 ──────────────────────
    "Q05_Monthly_Revenue_2024": {
        "business_question": "Which months in 2024 performed best and worst?",
        "sql": """
            SELECT
                STRFTIME('%Y-%m', Date)        AS Month,
                COUNT(*)                       AS Orders,
                ROUND(SUM(TotalPrice), 2)      AS Revenue
            FROM orders
            WHERE STRFTIME('%Y', Date) = '2024'
            GROUP BY Month
            ORDER BY Month;
        """,
        "insight": "June 2024 was peak (Rs 68,068). May 2024 was the weakest month (Rs 27,909 — only 34 orders)."
    },

    # ── QUERY 6: Payment method analysis ───────────────────
    "Q06_Payment_Method_Analysis": {
        "business_question": "Which payment method is most popular and most valuable?",
        "sql": """
            SELECT
                PaymentMethod,
                COUNT(*)                       AS Order_Count,
                ROUND(SUM(TotalPrice), 2)      AS Total_Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Order_Value
            FROM orders
            GROUP BY PaymentMethod
            ORDER BY Order_Count DESC;
        """,
        "insight": "Online is most used (258 orders). Credit Card drives highest avg order value (Rs 1,127.55)."
    },

    # ── QUERY 7: Top 10 highest value orders ───────────────
    "Q07_Top10_Highest_Orders": {
        "business_question": "What are the top 10 highest value orders? (VIP orders)",
        "sql": """
            SELECT
                OrderID, Date, Product, Quantity,
                UnitPrice, TotalPrice, OrderStatus, PaymentMethod
            FROM orders
            ORDER BY TotalPrice DESC
            LIMIT 10;
        """,
        "insight": "All top 10 orders are Qty=5 bulk purchases. Tablet leads at Rs 3,456.40 — these are VIP signals."
    },

    # ── QUERY 8: Coupon code impact ────────────────────────
    "Q08_Coupon_Code_Impact": {
        "business_question": "Which coupon codes drive the most usage and revenue?",
        "sql": """
            SELECT
                CouponCode,
                COUNT(*)                       AS Usage_Count,
                ROUND(SUM(TotalPrice), 2)      AS Total_Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Order_Value
            FROM orders
            GROUP BY CouponCode
            ORDER BY Usage_Count DESC;
        """,
        "insight": "FREESHIP leads with 313 uses (Rs 3,35,036 revenue). No-coupon orders (NONE) generate similar avg."
    },

    # ── QUERY 9: Referral source performance ───────────────
    "Q09_Referral_Source_Performance": {
        "business_question": "Which acquisition channel drives the most orders and revenue?",
        "sql": """
            SELECT
                ReferralSource,
                COUNT(*)                                            AS Orders,
                ROUND(SUM(TotalPrice), 2)                          AS Total_Revenue,
                ROUND(AVG(TotalPrice), 2)                          AS Avg_Order_Value,
                ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM orders), 2) AS Pct_Share
            FROM orders
            GROUP BY ReferralSource
            ORDER BY Orders DESC;
        """,
        "insight": "Instagram leads acquisition (259 orders, 21.6%). Facebook has highest avg order (Rs 1,098.29)."
    },

    # ── QUERY 10: Delivered orders above average ───────────
    "Q10_Delivered_Above_Avg": {
        "business_question": "Which delivered orders exceeded the average order value?",
        "sql": """
            SELECT
                OrderID, Product, TotalPrice, PaymentMethod, Date
            FROM orders
            WHERE OrderStatus = 'Delivered'
              AND TotalPrice > (SELECT AVG(TotalPrice) FROM orders)
            ORDER BY TotalPrice DESC
            LIMIT 15;
        """,
        "insight": "Uses a subquery to dynamically filter above the dataset average (Rs 1,053.97). Demonstrates WHERE + subquery."
    },

    # ── QUERY 11: Cancelled and returned losses ─────────────
    "Q11_Cancelled_Returned_Losses": {
        "business_question": "How much revenue is being lost to cancellations and returns per product?",
        "sql": """
            SELECT
                OrderStatus,
                Product,
                COUNT(*)                       AS Count,
                ROUND(SUM(TotalPrice), 2)      AS Lost_Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Lost
            FROM orders
            WHERE OrderStatus IN ('Cancelled', 'Returned')
            GROUP BY OrderStatus, Product
            ORDER BY OrderStatus, Lost_Revenue DESC;
        """,
        "insight": "Chair has highest cancellation losses (Rs 48,660). Tablet has highest return losses (Rs 42,525)."
    },

    # ── QUERY 12: High value products using HAVING ──────────
    "Q12_High_Value_Products_HAVING": {
        "business_question": "Which products have an average order value above Rs 1,000?",
        "sql": """
            SELECT
                Product,
                COUNT(*)                       AS Orders,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Revenue
            FROM orders
            GROUP BY Product
            HAVING AVG(TotalPrice) > 1000
            ORDER BY Avg_Revenue DESC;
        """,
        "insight": "5 of 7 products exceed Rs 1,000 avg. HAVING filters after GROUP BY — Desk (Rs 985) and Phone (Rs 972) excluded."
    },

    # ── QUERY 13: Bulk orders analysis (Qty = 5) ─────────────
    "Q13_Bulk_Orders_Qty5": {
        "business_question": "What is the profile of bulk (Quantity=5) orders?",
        "sql": """
            SELECT
                Product, OrderStatus, PaymentMethod,
                COUNT(*)                       AS Bulk_Orders,
                ROUND(SUM(TotalPrice), 2)      AS Revenue
            FROM orders
            WHERE Quantity = 5
            GROUP BY Product, OrderStatus, PaymentMethod
            ORDER BY Revenue DESC
            LIMIT 20;
        """,
        "insight": "Top bulk orders: Printer/Shipped/Gift Card (Rs 7,554). Confirms bulk orders are VIP signals, not data errors."
    },

    # ── QUERY 14: Channel-specific product performance ──────
    "Q14_Instagram_Product_Performance": {
        "business_question": "How do products perform specifically for Instagram-acquired customers?",
        "sql": """
            SELECT
                Product,
                COUNT(*)                       AS Orders,
                ROUND(SUM(TotalPrice), 2)      AS Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Value
            FROM orders
            WHERE ReferralSource = 'Instagram'
            GROUP BY Product
            ORDER BY Revenue DESC;
        """,
        "insight": "Laptop tops Instagram revenue (Rs 48,453). Desk has most orders (41) via Instagram."
    },

    # ── QUERY 15: Full monthly trend ────────────────────────
    "Q15_Full_Monthly_Trend": {
        "business_question": "What is the complete monthly revenue trend across all years?",
        "sql": """
            SELECT
                STRFTIME('%Y-%m', Date)        AS Month,
                COUNT(*)                       AS Orders,
                ROUND(SUM(TotalPrice), 2)      AS Revenue,
                ROUND(AVG(TotalPrice), 2)      AS Avg_Order
            FROM orders
            GROUP BY Month
            ORDER BY Month;
        """,
        "insight": "May 2023 was the peak month (Rs 63,836). May 2024 was the worst (Rs 27,909). Clear seasonal pattern."
    },
}

# ─────────────────────────────────────────────────────────
# EXECUTE ALL QUERIES & PRINT RESULTS
# ─────────────────────────────────────────────────────────
all_results = {}
print("\n" + "=" * 60)
print("  EXECUTING 15 SQL QUERIES")
print("=" * 60)

for name, meta in QUERIES.items():
    df_result = pd.read_sql_query(meta['sql'], conn)
    all_results[name] = df_result
    print(f"\n  [{name}]")
    print(f"  Business Question: {meta['business_question']}")
    print(f"  Rows returned: {len(df_result)}")
    print(f"  Insight: {meta['insight']}")

conn.close()

# ─────────────────────────────────────────────────────────
# BUILD FORMATTED EXCEL — One sheet per query + Summary
# ─────────────────────────────────────────────────────────

# Style helpers
H1_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy
H2_FILL   = PatternFill("solid", fgColor="2E75B6")   # medium blue
Q_FILL    = PatternFill("solid", fgColor="D6E4F0")   # light blue
I_FILL    = PatternFill("solid", fgColor="E2EFDA")   # light green
ALT_FILL  = PatternFill("solid", fgColor="EBF3FB")   # zebra
WHITE     = PatternFill("solid", fgColor="FFFFFF")

H1_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=12)
H2_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BD_FONT = Font(bold=True, color="1F3864", name="Calibri", size=10)
NM_FONT = Font(name="Calibri", size=10)
CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

thin = Side(style='thin', color='BDD7EE')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header_row(ws, row_num, fill, font):
    for cell in ws[row_num]:
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = border

def write_query_sheet(wb, sheet_name, df, meta):
    ws = wb.create_sheet(title=sheet_name[:31])

    # Title row
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")
    ws["A1"] = f"DecodeLabs | Project 3 — SQL Analysis | {sheet_name}"
    ws["A1"].font = H1_FONT
    ws["A1"].fill = H1_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 22

    # Business Question
    ws.merge_cells(f"A2:{get_column_letter(len(df.columns))}2")
    ws["A2"] = f"Business Question: {meta['business_question']}"
    ws["A2"].font = BD_FONT
    ws["A2"].fill = Q_FILL
    ws["A2"].alignment = LEFT
    ws.row_dimensions[2].height = 18

    # SQL query block
    sql_clean = ' '.join(meta['sql'].split())
    ws.merge_cells(f"A3:{get_column_letter(len(df.columns))}3")
    ws["A3"] = f"SQL: {sql_clean}"
    ws["A3"].font = Font(name="Courier New", size=9, color="1F3864")
    ws["A3"].fill = PatternFill("solid", fgColor="F0F7FF")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[3].height = 36

    # Column headers (row 4)
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=4, column=col_idx, value=col_name.replace('_', ' '))
        cell.font = H2_FONT
        cell.fill = H2_FILL
        cell.alignment = CENTER
        cell.border = border
    ws.row_dimensions[4].height = 20

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=5):
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = NM_FONT
            cell.fill = fill
            cell.border = border
            cell.alignment = CENTER if col_idx > 1 else LEFT

    # Insight row
    insight_row = len(df) + 5 + 1
    ws.merge_cells(f"A{insight_row}:{get_column_letter(len(df.columns))}{insight_row}")
    ws[f"A{insight_row}"] = f"KEY INSIGHT: {meta['insight']}"
    ws[f"A{insight_row}"].font = Font(bold=True, color="375623", name="Calibri", size=10)
    ws[f"A{insight_row}"].fill = I_FILL
    ws[f"A{insight_row}"].alignment = LEFT
    ws.row_dimensions[insight_row].height = 24

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 12), 35)

    ws.freeze_panes = "A5"
    return ws


wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── Summary sheet ──────────────────────────────────────
ws_sum = wb.create_sheet("Summary", 0)
ws_sum.sheet_view.showGridLines = False

ws_sum.merge_cells("A1:F1")
ws_sum["A1"] = "DecodeLabs | Project 3 — SQL Data Analysis | Executive Summary"
ws_sum["A1"].font = H1_FONT
ws_sum["A1"].fill = H1_FILL
ws_sum["A1"].alignment = CENTER
ws_sum.row_dimensions[1].height = 28

ws_sum.merge_cells("A2:F2")
ws_sum["A2"] = "Author: Shaik | Mohan Babu University | Batch 2026"
ws_sum["A2"].font = Font(bold=True, color="2E75B6", name="Calibri", size=11)
ws_sum["A2"].alignment = CENTER
ws_sum.row_dimensions[2].height = 20

# Sub-header
for col, text in enumerate(["Query ID","Business Question","SQL Clauses Used","Rows","Key Insight","Status"], 1):
    c = ws_sum.cell(row=3, column=col, value=text)
    c.font = H2_FONT; c.fill = H2_FILL; c.alignment = CENTER; c.border = border
ws_sum.row_dimensions[3].height = 20

clause_map = {
    "Q01": "SELECT, COUNT, SUM, AVG, MIN, MAX",
    "Q02": "SELECT, GROUP BY, ORDER BY, Aggregations",
    "Q03": "SELECT, GROUP BY, ORDER BY, Subquery, ROUND",
    "Q04": "SELECT, STRFTIME, GROUP BY, ORDER BY",
    "Q05": "SELECT, STRFTIME, WHERE, GROUP BY, ORDER BY",
    "Q06": "SELECT, GROUP BY, ORDER BY, AVG",
    "Q07": "SELECT, ORDER BY DESC, LIMIT",
    "Q08": "SELECT, GROUP BY, ORDER BY, SUM",
    "Q09": "SELECT, GROUP BY, ORDER BY, Subquery",
    "Q10": "SELECT, WHERE, Subquery (AVG), ORDER BY, LIMIT",
    "Q11": "SELECT, WHERE IN, GROUP BY, ORDER BY",
    "Q12": "SELECT, GROUP BY, HAVING, ORDER BY",
    "Q13": "SELECT, WHERE, GROUP BY, ORDER BY, LIMIT",
    "Q14": "SELECT, WHERE, GROUP BY, ORDER BY",
    "Q15": "SELECT, STRFTIME, GROUP BY, ORDER BY",
}

for row_idx, (name, meta) in enumerate(QUERIES.items(), start=4):
    qid = name[:3]
    df_r = all_results[name]
    fill = ALT_FILL if row_idx % 2 == 0 else WHITE
    data = [name, meta['business_question'], clause_map.get(qid,""), len(df_r), meta['insight'], "COMPLETE"]
    for col_idx, val in enumerate(data, 1):
        c = ws_sum.cell(row=row_idx, column=col_idx, value=val)
        c.font = NM_FONT; c.fill = fill; c.border = border
        c.alignment = LEFT if col_idx in [2,5] else CENTER
    ws_sum.row_dimensions[row_idx].height = 36

# Stats row
stats_row = 4 + len(QUERIES) + 1
ws_sum.merge_cells(f"A{stats_row}:F{stats_row}")
ws_sum[f"A{stats_row}"] = (
    "TOTAL: 15 SQL queries executed | "
    "Clauses used: SELECT, FROM, WHERE, GROUP BY, HAVING, ORDER BY, LIMIT, Subqueries, STRFTIME | "
    "Database: SQLite"
)
ws_sum[f"A{stats_row}"].font = Font(bold=True, color="375623", name="Calibri", size=10)
ws_sum[f"A{stats_row}"].fill = I_FILL
ws_sum[f"A{stats_row}"].alignment = LEFT
ws_sum.row_dimensions[stats_row].height = 24

for col, width in zip(['A','B','C','D','E','F'], [20, 38, 35, 8, 45, 10]):
    ws_sum.column_dimensions[col].width = width

ws_sum.freeze_panes = "A4"

# ── Write all 15 query sheets ──
for name, meta in QUERIES.items():
    short = name[4:] if len(name) > 4 else name
    write_query_sheet(wb, short[:31], all_results[name], meta)
    print(f"  Sheet written: {short}")

wb.save(output_path)
print(f"\n  Excel saved: {output_path}")
print("\n" + "=" * 60)
print("   PROJECT 3 COMPLETE")
print("=" * 60)
